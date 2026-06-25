import re
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from pypdf import PdfReader

st.set_page_config(page_title="HLD → NC Liste Generator", layout="wide")
st.title("HLD → NC Liste Generator")
st.write("Upload HLD PDF i NC Excel listu. Aplikacija izvuče osnovne HLD podatke i doda ih u sheet **AUTO_Extrakt**.")


def first_match(text: str, patterns, default=""):
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if m:
            return " ".join(m.group(1).strip().split())
    return default


def extract_pdf_text(pdf_file) -> str:
    reader = PdfReader(pdf_file)
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        pages.append(f"\n--- PAGE {i} ---\n{page_text}")
    return "\n".join(pages)


def extract_fields(text: str) -> dict:
    project = first_match(text, [r"Projekt(?:name)?\s*:?\s*([A-Z0-9_\-]+_[A-Za-zÄÖÜäöüß]+_[A-Za-zÄÖÜäöüß]+_[A-Z]{2})"])
    jira = first_match(text, [r"JIRA\s*:?\s*(ANDE\s*-?\s*\d+)"]).replace(" ", "")
    fibernode = first_match(text, [r"Fibernode\s*:?\s*([0-9A-Z]+-[0-9]+)", r"Bezeichnung FN alt\s*([0-9A-Z]+-[0-9]+)"])
    hub = first_match(text, [r"HUB\s+([A-Za-zÄÖÜäöüß\- ]+)", r"Hub\s*:?\s*([A-Za-zÄÖÜäöüß\- ]+)"])
    netz = first_match(text, [r"Netz\s*:?\s*([A-Za-zÄÖÜäöüß\- ]+)"])
    onkz = first_match(text, [r"ONKz\s*:?\s*(\d+)"])
    unlocode = first_match(text, [r"UNLOCODE\s*:?\s*([A-Z]{2}-[A-Z0-9]+)"])
    vrp = first_match(text, [r"VrP\s+([0-9]{4}-[0-9]{3}-[0-9]{3})", r"Standort DFN .*?\n\s*([0-9]{4}-[0-9]{3}-[0-9]{3})"])
    address = first_match(text, [r"(?:Standort DFN.*?\n[0-9]{4}-[0-9]{3}-[0-9]{3}\n)([^\n]+)"])
    postal_city = first_match(text, [r"\n\s*(\d{5}\s+[A-Za-zÄÖÜäöüß\- ]+)\s*\n"])
    gf_len = first_match(text, [r"GF-Länge\s*(\d+\s*m)", r"Länge\s*:?\s*(~?\s*\d+\s*m)"])
    dfn_ids = sorted(set(re.findall(r"UM-DF-\d+", text)))
    capex = first_match(text, [r"Zwischensumme\s*€?\s*([0-9\.]+)", r"Gesamtkosten\s*€?\s*([0-9\.]+)"])

    return {
        "Projektname": project,
        "JIRA": jira,
        "Fibernode": fibernode,
        "HUB": hub,
        "Netz/Ort": netz,
        "ONKz": onkz,
        "UNLOCODE": unlocode,
        "VrP / Standort DFN": vrp,
        "Adresse": address,
        "PLZ Ort": postal_city,
        "GF-Länge": gf_len,
        "UM-DF-IDs": ", ".join(dfn_ids),
        "Capex Zwischensumme": capex,
    }


def make_excel(template_file, fields: dict) -> bytes:
    wb = load_workbook(template_file)
    if "AUTO_Extrakt" in wb.sheetnames:
        del wb["AUTO_Extrakt"]
    ws = wb.create_sheet("AUTO_Extrakt", 0)
    ws["A1"] = "Feld"
    ws["B1"] = "Wert"
    ws["C1"] = "Quelle"
    for row, (key, value) in enumerate(fields.items(), start=2):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value="HLD PDF")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


pdf = st.file_uploader("1) HLD PDF", type=["pdf"])
xlsx = st.file_uploader("2) NC Excel lista", type=["xlsx"])

if pdf and xlsx:
    if st.button("Analiziraj PDF i generiraj NC listu"):
        with st.spinner("Čitam PDF i pripremam Excel..."):
            text = extract_pdf_text(pdf)
            fields = extract_fields(text)
            excel_bytes = make_excel(xlsx, fields)

        st.success("Gotovo. Provjeri pronađene podatke i preuzmi Excel.")
        st.subheader("Pronađeni podaci")
        st.dataframe([{"Feld": k, "Wert": v} for k, v in fields.items()], use_container_width=True)
        st.download_button(
            "Download popunjene NC liste",
            data=excel_bytes,
            file_name="NC-Liste_AUTO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Uploaduj oba fajla pa klikni gumb za generiranje.")
